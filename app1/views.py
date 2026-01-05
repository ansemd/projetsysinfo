from django.shortcuts import render, get_object_or_404, redirect
from django.contrib import messages
from django.db.models import Q, Count, Sum
from .models import Client, Chauffeur


# ========== LISTE DES CLIENTS (PAGE PRINCIPALE) ==========
def liste_clients(request):
    """
    Page principale : Liste de tous les clients avec recherche et statistiques
    """
    search = request.GET.get('search', '')
    
    if search:
        # Recherche par nom, prénom, téléphone, email, ville ou wilaya
        clients = Client.objects.filter(
            Q(nom__icontains=search) | 
            Q(prenom__icontains=search) |
            Q(telephone__icontains=search) |
            Q(email__icontains=search) |
            Q(ville__icontains=search) |
            Q(wilaya__icontains=search)
        ).order_by('-date_inscription')
    else:
        clients = Client.objects.all().order_by('-date_inscription')
    
    # Statistiques globales
    stats = {
        'total_clients': Client.objects.count(),
        'clients_avec_solde': Client.objects.filter(solde__gt=0).count(),
        'solde_total': Client.objects.aggregate(Sum('solde'))['solde__sum'] or 0,
    }
    
    return render(request, 'clients/liste.html', {
        'clients': clients,
        'search': search,
        'stats': stats,
    })

# ========== DÉTAILS D'UN CLIENT ==========
def detail_client(request, client_id):
    """
    Affiche tous les détails d'un client + ses expéditions
    """
    client = get_object_or_404(Client, id=client_id)
    
    # Récupérer toutes les expéditions du client
    expeditions = client.expedition_set.all().order_by('-date_creation')
    
    # Statistiques du client
    stats_client = {
        'total_expeditions': expeditions.count(),
        'expeditions_livrees': expeditions.filter(statut='LIVRE').count(),
        'expeditions_en_cours': expeditions.filter(statut='EN_TRANSIT').count(),
        'expeditions_en_attente': expeditions.filter(statut='EN_ATTENTE').count(),
        'total_depense': expeditions.aggregate(Sum('montant_total'))['montant_total__sum'] or 0,
    }
    
    # Récupérer les factures du client
    factures = client.factures.all().order_by('-date_creation')[:5]  # Les 5 dernières
    
    return render(request, 'clients/detail.html', {
        'client': client,
        'expeditions': expeditions,
        'stats_client': stats_client,
        'factures': factures,
    })

# ========== CRÉER UN CLIENT ==========
def creer_client(request):
    """
    Formulaire de création d'un nouveau client
    """
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            client = Client(
                nom=request.POST['nom'],
                prenom=request.POST['prenom'],
                date_naissance=request.POST['date_naissance'],
                telephone=request.POST['telephone'],
                email=request.POST.get('email', ''),
                adresse=request.POST.get('adresse', ''),
                ville=request.POST.get('ville', ''),
                wilaya=request.POST.get('wilaya', ''),
                solde=request.POST.get('solde', 0.00),
                remarques=request.POST.get('remarques', ''),
            )
            client.save()
            
            messages.success(request, f'Client {client.prenom} {client.nom} créé avec succès!')
            return redirect('detail_client', client_id=client.id)
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du client: {str(e)}')
    
    return render(request, 'clients/creer.html')

# ========== MODIFIER UN CLIENT ==========
def modifier_client(request, client_id):
    """
    Formulaire de modification d'un client existant
    """
    client = get_object_or_404(Client, id=client_id)
    
    if request.method == 'POST':
        try:
            # Mettre à jour les données du client
            client.nom = request.POST['nom']
            client.prenom = request.POST['prenom']
            client.date_naissance = request.POST['date_naissance']
            client.telephone = request.POST['telephone']
            client.email = request.POST.get('email', '')
            client.adresse = request.POST.get('adresse', '')
            client.ville = request.POST.get('ville', '')
            client.wilaya = request.POST.get('wilaya', '')
            client.solde = request.POST.get('solde', 0.00)
            client.remarques = request.POST.get('remarques', '')
            
            # La date_modification sera mise à jour automatiquement (auto_now=True)
            client.save()
            
            messages.success(request, f'Client {client.prenom} {client.nom} modifié avec succès!')
            return redirect('detail_client', client_id=client.id)
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification: {str(e)}')
    
    return render(request, 'clients/modifier.html', {
        'client': client,
    })

# ========== SUPPRIMER UN CLIENT ==========
def supprimer_client(request, client_id):
    """
    Suppression d'un client (avec confirmation)
    """
    client = get_object_or_404(Client, id=client_id)
    
    if request.method == 'POST':
        try:
            nom_complet = f"{client.prenom} {client.nom}"
            client.delete()
            messages.success(request, f'Client {nom_complet} supprimé avec succès!')
            return redirect('liste_clients')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la suppression: {str(e)}')
            return redirect('detail_client', client_id=client_id)
    
    # Si GET, afficher page de confirmation
    return render(request, 'clients/supprimer.html', {
        'client': client,
    })


# ========== LISTE DES CHAUFFEURS (PAGE PRINCIPALE) ==========
def liste_chauffeurs(request):
    """
    Page principale : Liste de tous les chauffeurs avec recherche et filtrage par statut
    """
    search = request.GET.get('search', '')
    filtre_statut = request.GET.get('statut', '')  # Filtre par statut
    
    # Base queryset
    chauffeurs = Chauffeur.objects.all()
    
    # Recherche
    if search:
        chauffeurs = chauffeurs.filter(
            Q(nom__icontains=search) | 
            Q(prenom__icontains=search) |
            Q(numero_permis__icontains=search) |
            Q(telephone__icontains=search) |
            Q(ville__icontains=search) |
            Q(wilaya__icontains=search)
        )
    
    # Filtrage par statut
    if filtre_statut:
        chauffeurs = chauffeurs.filter(statut_disponibilite=filtre_statut)
    
    chauffeurs = chauffeurs.order_by('-date_embauche')
    
    # Statistiques globales
    stats = {
        'total_chauffeurs': Chauffeur.objects.count(),
        'disponibles': Chauffeur.objects.filter(statut_disponibilite='DISPONIBLE').count(),
        'en_tournee': Chauffeur.objects.filter(statut_disponibilite='EN_TOURNEE').count(),
        'en_conge': Chauffeur.objects.filter(statut_disponibilite='CONGE').count(),
    }
    
    # Liste des statuts pour le filtre
    statuts = [
        ('DISPONIBLE', 'Disponible'),
        ('EN_TOURNEE', 'En tournée'),
        ('CONGE', 'En congé'),
        ('MALADIE', 'Maladie'),
        ('AUTRE', 'Autre raison'),
    ]
    
    return render(request, 'chauffeurs/liste.html', {
        'chauffeurs': chauffeurs,
        'search': search,
        'filtre_statut': filtre_statut,
        'statuts': statuts,
        'stats': stats,
    })

# ========== MODIFIER STATUT CHAUFFEUR (AJAX) ==========
def modifier_statut_chauffeur(request, chauffeur_id):
    """
    Modifie uniquement le statut d'un chauffeur (appelé depuis la liste)
    """
    if request.method == 'POST':
        chauffeur = get_object_or_404(Chauffeur, id=chauffeur_id)
        nouveau_statut = request.POST.get('statut')
        
        if nouveau_statut in ['DISPONIBLE', 'EN_TOURNEE', 'CONGE', 'MALADIE', 'AUTRE']:
            chauffeur.statut_disponibilite = nouveau_statut
            chauffeur.save()
            messages.success(request, f'Statut de {chauffeur.prenom} {chauffeur.nom} modifié avec succès!')
        else:
            messages.error(request, 'Statut invalide')
    
    return redirect('liste_chauffeurs')

# ========== DÉTAILS D'UN CHAUFFEUR ==========
def detail_chauffeur(request, chauffeur_id):
    """
    Affiche tous les détails d'un chauffeur + sa tournée actuelle (EN_COURS ou PREVUE)
    """
    chauffeur = get_object_or_404(Chauffeur, id=chauffeur_id)
    
    # Récupérer SEULEMENT la tournée actuelle (EN_COURS ou PREVUE)
    tournee_actuelle = chauffeur.tournee_set.filter(
        statut__in=['EN_COURS', 'PREVUE']
    ).order_by('-date_creation').first()
    
    # Statistiques du chauffeur (toutes les tournées)
    all_tournees = chauffeur.tournee_set.all()
    stats_chauffeur = {
        'total_tournees': all_tournees.count(),
        'tournees_prevues': all_tournees.filter(statut='PREVUE').count(),
        'tournees_en_cours': all_tournees.filter(statut='EN_COURS').count(),
        'tournees_terminees': all_tournees.filter(statut='TERMINEE').count(),
    }
    
    return render(request, 'chauffeurs/detail.html', {
        'chauffeur': chauffeur,
        'tournee_actuelle': tournee_actuelle,
        'stats_chauffeur': stats_chauffeur,
    })

# ========== CRÉER UN CHAUFFEUR ==========
def creer_chauffeur(request):
    """
    Formulaire de création d'un nouveau chauffeur
    """
    if request.method == 'POST':
        try:
            from datetime import datetime
            
            # Récupérer et valider les dates
            date_obtention = datetime.strptime(request.POST['date_obtention_permis'], '%Y-%m-%d').date()
            date_expiration = datetime.strptime(request.POST['date_expiration_permis'], '%Y-%m-%d').date()
            
            # VALIDATION : date obtention doit être AVANT date expiration
            if date_obtention >= date_expiration:
                messages.error(
                    request, 
                    'Erreur : La date d\'obtention du permis doit être antérieure à la date d\'expiration !'
                )
                return render(request, 'chauffeurs/creer.html')
            
            chauffeur = Chauffeur(
                nom=request.POST['nom'],
                prenom=request.POST['prenom'],
                date_naissance=request.POST['date_naissance'],
                telephone=request.POST['telephone'],
                email=request.POST.get('email', ''),
                adresse=request.POST.get('adresse', ''),
                ville=request.POST.get('ville', ''),
                wilaya=request.POST.get('wilaya', ''),
                numero_permis=request.POST['numero_permis'],
                date_obtention_permis=date_obtention,
                date_expiration_permis=date_expiration,
                date_embauche=request.POST['date_embauche'],
                salaire=request.POST.get('salaire', None),
                statut_disponibilite=request.POST.get('statut_disponibilite', 'DISPONIBLE'),
                remarques=request.POST.get('remarques', ''),
            )
            chauffeur.save()
            
            messages.success(request, f'Chauffeur {chauffeur.prenom} {chauffeur.nom} créé avec succès!')
            return redirect('detail_chauffeur', chauffeur_id=chauffeur.id)
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création du chauffeur: {str(e)}')
    
    return render(request, 'chauffeurs/creer.html')

# ========== MODIFIER UN CHAUFFEUR ==========
def modifier_chauffeur(request, chauffeur_id):
    """
    Formulaire de modification d'un chauffeur existant
    """
    chauffeur = get_object_or_404(Chauffeur, id=chauffeur_id)
    
    if request.method == 'POST':
        try:
            from datetime import datetime
            
            # Récupérer et valider les dates
            date_obtention = datetime.strptime(request.POST['date_obtention_permis'], '%Y-%m-%d').date()
            date_expiration = datetime.strptime(request.POST['date_expiration_permis'], '%Y-%m-%d').date()
            
            # VALIDATION : date obtention doit être AVANT date expiration
            if date_obtention >= date_expiration:
                messages.error(
                    request, 
                    'Erreur : La date d\'obtention du permis doit être antérieure à la date d\'expiration !'
                )
                return render(request, 'chauffeurs/modifier.html', {'chauffeur': chauffeur})
            
            chauffeur.nom = request.POST['nom']
            chauffeur.prenom = request.POST['prenom']
            chauffeur.date_naissance = request.POST['date_naissance']
            chauffeur.telephone = request.POST['telephone']
            chauffeur.email = request.POST.get('email', '')
            chauffeur.adresse = request.POST.get('adresse', '')
            chauffeur.ville = request.POST.get('ville', '')
            chauffeur.wilaya = request.POST.get('wilaya', '')
            chauffeur.numero_permis = request.POST['numero_permis']
            chauffeur.date_obtention_permis = date_obtention
            chauffeur.date_expiration_permis = date_expiration
            chauffeur.date_embauche = request.POST['date_embauche']
            chauffeur.salaire = request.POST.get('salaire', None)
            chauffeur.statut_disponibilite = request.POST.get('statut_disponibilite', 'DISPONIBLE')
            chauffeur.remarques = request.POST.get('remarques', '')
            
            chauffeur.save()
            
            messages.success(request, f'Chauffeur {chauffeur.prenom} {chauffeur.nom} modifié avec succès!')
            return redirect('detail_chauffeur', chauffeur_id=chauffeur.id)
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification: {str(e)}')
    
    return render(request, 'chauffeurs/modifier.html', {
        'chauffeur': chauffeur,
    })

# ========== SUPPRIMER UN CHAUFFEUR ==========
def supprimer_chauffeur(request, chauffeur_id):
    """
    Suppression d'un chauffeur (avec confirmation)
    """
    chauffeur = get_object_or_404(Chauffeur, id=chauffeur_id)
    
    if request.method == 'POST':
        try:
            nom_complet = f"{chauffeur.prenom} {chauffeur.nom}"
            chauffeur.delete()
            messages.success(request, f'Chauffeur {nom_complet} supprimé avec succès!')
            return redirect('liste_chauffeurs')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la suppression: {str(e)}')
            return redirect('detail_chauffeur', chauffeur_id=chauffeur_id)
    
    return render(request, 'chauffeurs/supprimer.html', {
        'chauffeur': chauffeur,
    })


from django.core.exceptions import ValidationError
from datetime import datetime, timedelta
from .models import Incident, Reclamation, HistoriqueReclamation, Expedition, Tournee, Client, Facture
from .utils import IncidentService, ReclamationService, ExpeditionService

# ========== VUES INCIDENTS ==========
def liste_incidents(request):
    """
    Page principale : Liste de tous les incidents avec recherche et filtres
    """
    search = request.GET.get('search', '')
    type_incident = request.GET.get('type_incident', '')
    severite = request.GET.get('severite', '')
    statut = request.GET.get('statut', '')
    
    # Base queryset
    incidents = Incident.objects.all()
    
    # Recherche
    if search:
        incidents = incidents.filter(
            Q(numero_incident__icontains=search) |
            Q(titre__icontains=search) |
            Q(description__icontains=search) |
            Q(signale_par__icontains=search) |
            Q(lieu_incident__icontains=search)
        )
    
    # Filtres
    if type_incident:
        incidents = incidents.filter(type_incident=type_incident)
    if severite:
        incidents = incidents.filter(severite=severite)
    if statut:
        incidents = incidents.filter(statut=statut)
    
    incidents = incidents.order_by('-date_heure_incident')
    
    # Statistiques globales
    stats = {
        'total': incidents.count(),
        'signales': Incident.objects.filter(statut='SIGNALE').count(),
        'en_cours': Incident.objects.filter(statut='EN_COURS').count(),
        'resolus': Incident.objects.filter(statut='RESOLU').count(),
        'clos': Incident.objects.filter(statut='CLOS').count(),
        'critiques': Incident.objects.filter(severite='CRITIQUE').count(),
        'eleves': Incident.objects.filter(severite='ELEVEE').count(),
    }
    
    # Choix pour les filtres
    types = [
        ('RETARD', 'Retard de livraison'),
        ('PERTE', 'Perte de colis'),
        ('ENDOMMAGEMENT', 'Endommagement'),
        ('PROBLEME_TECHNIQUE', 'Problème technique véhicule'),
        ('ACCIDENT', 'Accident'),
        ('REFUS_DESTINATAIRE', 'Refus du destinataire'),
        ('ADRESSE_INCORRECTE', 'Adresse incorrecte'),
        ('DESTINATAIRE_ABSENT', 'Destinataire absent'),
        ('AUTRE', 'Autre'),
    ]
    
    severites = [
        ('FAIBLE', 'Faible'),
        ('MOYENNE', 'Moyenne'),
        ('ELEVEE', 'Élevée'),
        ('CRITIQUE', 'Critique'),
    ]
    
    statuts = [
        ('SIGNALE', 'Signalé'),
        ('EN_COURS', 'En cours de traitement'),
        ('RESOLU', 'Résolu'),
        ('CLOS', 'Clôturé'),
    ]
    
    return render(request, 'incidents/liste.html', {
        'incidents': incidents,
        'search': search,
        'type_incident': type_incident,
        'severite': severite,
        'statut': statut,
        'types': types,
        'severites': severites,
        'statuts': statuts,
        'stats': stats,
    })


def detail_incident(request, incident_id):
    """
    Affiche tous les détails d'un incident
    """
    incident = get_object_or_404(Incident, id=incident_id)
    
    return render(request, 'incidents/detail.html', {
        'incident': incident,
    })


def creer_incident(request):
    """
    Formulaire de création d'incident avec annulation automatique d'expédition si nécessaire
    """
    if request.method == 'POST':
        try:
            # Récupérer les données du formulaire
            type_incident = request.POST.get('type_incident')
            titre = request.POST.get('titre')
            description = request.POST.get('description')
            date_heure = request.POST.get('date_heure_incident')
            lieu = request.POST.get('lieu_incident', '')
            signale_par = request.POST.get('signale_par')
            cout = request.POST.get('cout_estime', 0)
            
            # Récupérer expédition OU tournée (un seul des deux)
            expedition_id = request.POST.get('expedition_id')
            tournee_id = request.POST.get('tournee_id')

            document1 = request.FILES.get('document1')
            document2 = request.FILES.get('document2')
            photo1 = request.FILES.get('photo1')
            photo2 = request.FILES.get('photo2')

            
            # Validation : un seul des deux
            if expedition_id and tournee_id:
                messages.error(request, 'Un incident ne peut être lié qu\'à une expédition OU une tournée, pas les deux!')
                return render(request, 'incidents/creer.html', {
                    'types': Incident._meta.get_field('type_incident').choices,
                    'expeditions': Expedition.objects.all(),
                    'tournees': Tournee.objects.filter(statut__in=['PREVUE', 'EN_COURS']),
                })
            
            if not expedition_id and not tournee_id:
                messages.error(request, 'Un incident doit être lié à une expédition OU une tournée!')
                return render(request, 'incidents/creer.html', {
                    'types': Incident._meta.get_field('type_incident').choices,
                    'expeditions': Expedition.objects.all(),
                    'tournees': Tournee.objects.filter(statut__in=['PREVUE', 'EN_COURS']),
                })
            
            # Créer l'incident
            incident = Incident.objects.create(
                type_incident=type_incident,
                titre=titre,
                description=description,
                date_heure_incident=date_heure,
                lieu_incident=lieu,
                signale_par=signale_par,
                cout_estime=cout,
                expedition_id=expedition_id if expedition_id else None,
                tournee_id=tournee_id if tournee_id else None,
                document1=document1,
                document2=document2,
                photo1=photo1,
                photo2=photo2,
            )
            
            if type_incident in ['PERTE', 'ENDOMMAGEMENT', 'PROBLEME_TECHNIQUE', 'ACCIDENT'] and expedition_id:
                try:
                    expedition = Expedition.objects.get(id=expedition_id)
                    
                    # Vérifier si l'expédition peut être annulée
                    if expedition.statut in ['EN_ATTENTE', 'COLIS_CREE']:
                        # Annuler l'expédition
                        ExpeditionService.annuler_expedition(expedition)
                        messages.warning(
                            request,
                            f"Incident grave détecté : L'expédition {expedition.get_numero_expedition()} "
                            f"a été automatiquement annulée et le client remboursé."
                        )
                    else:
                        messages.info(
                            request,
                            f"L'expédition ne peut pas être annulée automatiquement car elle est en {expedition.get_statut_display()}. "
                            f"Veuillez la traiter manuellement."
                        )
                except Exception as e:
                    messages.error(request, f"Erreur lors de l'annulation automatique : {str(e)}")
            
            # Les emails sont envoyés automatiquement via le signal save() !
            messages.success(
                request,
                f"Incident {incident.numero_incident} créé avec succès ! "
                f"Alertes envoyées par email."
            )
            return redirect('detail_incident', incident_id=incident.id)
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création de l\'incident: {str(e)}')
    
    # GET : afficher le formulaire
    types = Incident._meta.get_field('type_incident').choices
    expeditions = Expedition.objects.filter(statut__in=['EN_ATTENTE', 'EN_TRANSIT', 'COLIS_CREE'])
    tournees = Tournee.objects.filter(statut__in=['PREVUE', 'EN_COURS'])
    
    return render(request, 'incidents/creer.html', {
        'types': types,
        'expeditions': expeditions,
        'tournees': tournees,
    })


def modifier_incident(request, incident_id):
    """
    Formulaire de modification d'un incident existant
    """
    incident = get_object_or_404(Incident, id=incident_id)
    
    if request.method == 'POST':
        try:
            # Mettre à jour les données modifiables
            incident.titre = request.POST.get('titre')
            incident.description = request.POST.get('description')
            incident.lieu_incident = request.POST.get('lieu_incident', '')
            incident.cout_estime = request.POST.get('cout_estime', 0)
            incident.severite = request.POST.get('severite')
            incident.statut = request.POST.get('statut')
            incident.agent_responsable = request.POST.get('agent_responsable', '')
            incident.actions_entreprises = request.POST.get('actions_entreprises', '')
            incident.remarques = request.POST.get('remarques', '')
            incident.document1 = request.FILES.get('document1', incident.document1)
            incident.document2 = request.FILES.get('document2', incident.document2)
            incident.photo1 = request.FILES.get('photo1', incident.photo1)
            incident.photo2 = request.FILES.get('photo2', incident.photo2)

            incident.save()
            
            messages.success(request, f'Incident {incident.numero_incident} modifié avec succès!')
            return redirect('detail_incident', incident_id=incident.id)
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification: {str(e)}')
    
    severites = Incident._meta.get_field('severite').choices
    statuts = Incident._meta.get_field('statut').choices
    
    return render(request, 'incidents/modifier.html', {
        'incident': incident,
        'severites': severites,
        'statuts': statuts,
    })


def resoudre_incident(request, incident_id):
    """
    Marquer un incident comme résolu
    """
    incident = get_object_or_404(Incident, id=incident_id)
    
    if request.method == 'POST':
        try:
            solution = request.POST.get('solution')
            agent = request.POST.get('agent', 'Agent')
            
            IncidentService.resoudre_incident(incident, solution, agent)
            
            messages.success(request, f'Incident {incident.numero_incident} résolu avec succès!')
            return redirect('detail_incident', incident_id=incident.id)
            
        except Exception as e:
            messages.error(request, f'Erreur: {str(e)}')
    
    return render(request, 'incidents/resoudre.html', {
        'incident': incident,
    })


def cloturer_incident(request, incident_id):
    """
    Clôturer définitivement un incident
    """
    incident = get_object_or_404(Incident, id=incident_id)
    
    if request.method == 'POST':
        try:
            IncidentService.cloturer_incident(incident)
            messages.success(request, f'Incident {incident.numero_incident} clôturé!')
            return redirect('liste_incidents')
            
        except ValidationError as e:
            messages.error(request, str(e))
            return redirect('detail_incident', incident_id=incident.id)
    
    return render(request, 'incidents/cloturer.html', {
        'incident': incident,
    })


def supprimer_incident(request, incident_id):
    """
    Suppression d'un incident (avec confirmation)
    """
    incident = get_object_or_404(Incident, id=incident_id)
    
    if request.method == 'POST':
        try:
            numero = incident.numero_incident
            incident.delete()
            messages.success(request, f'Incident {numero} supprimé avec succès!')
            return redirect('liste_incidents')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la suppression: {str(e)}')
            return redirect('detail_incident', incident_id=incident_id)
    
    return render(request, 'incidents/supprimer.html', {
        'incident': incident,
    })


# ========== VUES RÉCLAMATIONS ==========

def liste_reclamations(request):
    """
    Page principale : Liste de toutes les réclamations avec recherche et filtres
    """
    search = request.GET.get('search', '')
    type_reclamation = request.GET.get('type_reclamation', '')
    nature = request.GET.get('nature', '')
    priorite = request.GET.get('priorite', '')
    statut = request.GET.get('statut', '')
    client_id = request.GET.get('client_id', '')
    
    # Base queryset
    reclamations = Reclamation.objects.all()
    
    # Recherche
    if search:
        reclamations = reclamations.filter(
            Q(numero_reclamation__icontains=search) |
            Q(objet__icontains=search) |
            Q(description__icontains=search) |
            Q(client__nom__icontains=search) |
            Q(client__prenom__icontains=search)
        )
    
    # Filtres
    if type_reclamation:
        reclamations = reclamations.filter(type_reclamation=type_reclamation)
    if nature:
        reclamations = reclamations.filter(nature=nature)
    if priorite:
        reclamations = reclamations.filter(priorite=priorite)
    if statut:
        reclamations = reclamations.filter(statut=statut)
    if client_id:
        reclamations = reclamations.filter(client_id=client_id)
    
    reclamations = reclamations.order_by('-date_creation')
    
    # Statistiques globales
    stats = {
        'total': reclamations.count(),
        'ouvertes': Reclamation.objects.filter(statut='OUVERTE').count(),
        'en_cours': Reclamation.objects.filter(statut='EN_COURS').count(),
        'resolues': Reclamation.objects.filter(statut='RESOLUE').count(),
        'closes': Reclamation.objects.filter(statut='CLOSE').count(),
        'urgentes': Reclamation.objects.filter(priorite='URGENTE').count(),
        'avec_compensation': Reclamation.objects.filter(compensation_accordee=True).count(),
    }
    
    # Choix pour les filtres
    types = Reclamation._meta.get_field('type_reclamation').choices
    natures = Reclamation._meta.get_field('nature').choices
    priorites = Reclamation._meta.get_field('priorite').choices
    statuts = Reclamation._meta.get_field('statut').choices
    clients = Client.objects.all().order_by('nom', 'prenom')
    
    return render(request, 'reclamations/liste.html', {
        'reclamations': reclamations,
        'search': search,
        'type_reclamation': type_reclamation,
        'nature': nature,
        'priorite': priorite,
        'statut': statut,
        'client_id': client_id,
        'types': types,
        'natures': natures,
        'priorites': priorites,
        'statuts': statuts,
        'clients': clients,
        'stats': stats,
    })


def detail_reclamation(request, reclamation_id):
    """
    Affiche tous les détails d'une réclamation avec son historique
    """
    reclamation = get_object_or_404(Reclamation, id=reclamation_id)
    historique = reclamation.historique.all()
    
    return render(request, 'reclamations/detail.html', {
        'reclamation': reclamation,
        'historique': historique,
    })


def creer_reclamation(request):
    """
    Formulaire de création de réclamation
    """
    if request.method == 'POST':
        try:
            # Récupérer les données
            client_id = request.POST.get('client_id')
            type_reclamation = request.POST.get('type_reclamation')
            nature = request.POST.get('nature')
            objet = request.POST.get('objet')
            description = request.POST.get('description')
            service_concerne = request.POST.get('service_concerne', None)
            expedition_ids = request.POST.getlist('expeditions')
            facture_id = request.POST.get('facture_id', None)
            piece1 = request.FILES.get('piece_jointe1')
            piece2 = request.FILES.get('piece_jointe2')


            # Créer la réclamation
            reclamation = Reclamation.objects.create(
                client_id=client_id,
                type_reclamation=type_reclamation,
                nature=nature,
                objet=objet,
                description=description,
                service_concerne=service_concerne if service_concerne else None,
                facture_id=facture_id if facture_id else None,
                piece_jointe1=piece1,
                piece_jointe2=piece2,
            )
            
            # Ajouter les expéditions
            if expedition_ids:
                reclamation.expeditions.set(expedition_ids)
            
            # Les emails sont envoyés automatiquement via le signal save() !
            messages.success(
                request,
                f"Réclamation {reclamation.numero_reclamation} créée avec succès ! "
                f"Confirmation envoyée au client par email."
            )
            return redirect('detail_reclamation', reclamation_id=reclamation.id)
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la création de la réclamation: {str(e)}')
    
    # GET : afficher le formulaire
    types = Reclamation._meta.get_field('type_reclamation').choices
    natures = Reclamation._meta.get_field('nature').choices
    services = Reclamation._meta.get_field('service_concerne').choices
    clients = Client.objects.all().order_by('nom', 'prenom')
    expeditions = Expedition.objects.all().order_by('-date_creation')
    factures = Facture.objects.exclude(statut='ANNULEE').order_by('-date_creation')
    
    return render(request, 'reclamations/creer.html', {
        'types': types,
        'natures': natures,
        'services': services,
        'clients': clients,
        'expeditions': expeditions,
        'factures': factures,
    })


def modifier_reclamation(request, reclamation_id):
    """
    Formulaire de modification d'une réclamation existante
    """
    reclamation = get_object_or_404(Reclamation, id=reclamation_id)
    
    if request.method == 'POST':
        try:
            # Mettre à jour les données modifiables
            reclamation.objet = request.POST.get('objet')
            reclamation.description = request.POST.get('description')
            reclamation.priorite = request.POST.get('priorite')
            reclamation.statut = request.POST.get('statut')
            reclamation.agent_responsable = request.POST.get('agent_responsable', '')
            reclamation.remarques = request.POST.get('remarques', '')
            
            # Mettre à jour les expéditions
            expedition_ids = request.POST.getlist('expeditions')
            if expedition_ids:
                reclamation.expeditions.set(expedition_ids)
            
            reclamation.save()
            
            messages.success(request, f'Réclamation {reclamation.numero_reclamation} modifiée avec succès!')
            return redirect('detail_reclamation', reclamation_id=reclamation.id)
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la modification: {str(e)}')
    
    priorites = Reclamation._meta.get_field('priorite').choices
    statuts = Reclamation._meta.get_field('statut').choices
    expeditions = Expedition.objects.all().order_by('-date_creation')
    
    return render(request, 'reclamations/modifier.html', {
        'reclamation': reclamation,
        'priorites': priorites,
        'statuts': statuts,
        'expeditions': expeditions,
    })


def assigner_reclamation(request, reclamation_id):
    """
    Assigner une réclamation à un agent
    """
    reclamation = get_object_or_404(Reclamation, id=reclamation_id)
    
    if request.method == 'POST':
        try:
            agent_nom = request.POST.get('agent_nom')
            
            ReclamationService.assigner_agent(reclamation, agent_nom)
            
            messages.success(request, f'Réclamation assignée à {agent_nom}')
            return redirect('detail_reclamation', reclamation_id=reclamation.id)
            
        except Exception as e:
            messages.error(request, f'Erreur: {str(e)}')
    
    return render(request, 'reclamations/assigner.html', {
        'reclamation': reclamation,
    })


def resoudre_reclamation(request, reclamation_id):
    """
    Résoudre une réclamation avec compensation éventuelle
    """
    reclamation = get_object_or_404(Reclamation, id=reclamation_id)
    
    if request.method == 'POST':
        try:
            auteur = request.POST.get('auteur', 'Agent')
            accorder_compensation = request.POST.get('compensation') == 'on'
            montant_compensation = float(request.POST.get('montant_compensation', 0))
            
            ReclamationService.resoudre_reclamation(
                reclamation,
                auteur,
                accorder_compensation,
                montant_compensation
            )
            
            # Email envoyé automatiquement !
            messages.success(
                request,
                f'Réclamation {reclamation.numero_reclamation} résolue ! '
                f'Email de confirmation envoyé au client.'
            )
            return redirect('detail_reclamation', reclamation_id=reclamation.id)
            
        except Exception as e:
            messages.error(request, f'Erreur: {str(e)}')
    
    return render(request, 'reclamations/resoudre.html', {
        'reclamation': reclamation,
    })


def cloturer_reclamation(request, reclamation_id):
    """
    Clôturer définitivement une réclamation
    """
    reclamation = get_object_or_404(Reclamation, id=reclamation_id)
    
    if request.method == 'POST':
        try:
            auteur = request.POST.get('auteur', 'Agent')
            ReclamationService.cloturer_reclamation(reclamation, auteur)
            
            messages.success(request, f'Réclamation {reclamation.numero_reclamation} clôturée!')
            return redirect('liste_reclamations')
            
        except ValidationError as e:
            messages.error(request, str(e))
            return redirect('detail_reclamation', reclamation_id=reclamation.id)
    
    return render(request, 'reclamations/cloturer.html', {
        'reclamation': reclamation,
    })


def supprimer_reclamation(request, reclamation_id):
    """
    Suppression d'une réclamation (avec confirmation)
    """
    reclamation = get_object_or_404(Reclamation, id=reclamation_id)
    
    if request.method == 'POST':
        try:
            numero = reclamation.numero_reclamation
            reclamation.delete()
            messages.success(request, f'Réclamation {numero} supprimée avec succès!')
            return redirect('liste_reclamations')
            
        except Exception as e:
            messages.error(request, f'Erreur lors de la suppression: {str(e)}')
            return redirect('detail_reclamation', reclamation_id=reclamation_id)
    
    return render(request, 'reclamations/supprimer.html', {
        'reclamation': reclamation,
    })

from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_http_methods
from datetime import datetime
from decimal import Decimal
import json
from app1.services.analytics_service import AnalyticsService
from app1.services.stats_service import StatsService
from app1.models import Expedition, Tournee, Client, Chauffeur

# ==================== TABLEAU DE BORD PRINCIPAL ====================

@login_required
def dashboard_analytics(request):
    """
    Vue principale du tableau de bord d'analyse
    """
    annee = int(request.GET.get('annee', datetime.now().year))
    
    # Récupérer les années disponibles depuis la base de données
    annees_disponibles = list( Expedition.objects.dates('date_creation', 'year') .values_list('date_creation__year', flat=True)
        .distinct().order_by('-date_creation__year') )
    
    # Si aucune donnée, utiliser une plage par défaut
    if not annees_disponibles:
        annees_disponibles = range(2020, datetime.now().year + 1)
    
    context = {'annee': annee, 'annees_disponibles': annees_disponibles,'titre': 'Tableau de Bord - Analyse et Statistiques'}
    
    return render(request, 'analytics/dashboard.html', context)


# ==================== ANALYSE COMMERCIALE ====================

@login_required
def analyse_commerciale(request):
    """
    Page d'analyse commerciale détaillée
    """
    annee = int(request.GET.get('annee', datetime.now().year))
    annees_disponibles = list( Expedition.objects.dates('date_creation', 'year') .values_list('date_creation__year', flat=True)
        .distinct() .order_by('-date_creation__year') )
    
    if not annees_disponibles:
        annees_disponibles = range(2020, datetime.now().year + 1)
    
    context = {'annee': annee,'annees_disponibles': annees_disponibles,'titre': 'Analyse Commerciale'}
    
    return render(request, 'analytics/commerciale.html', context)


@login_required
@require_http_methods(["GET"])
def api_evolution_expeditions(request):
    """
    API pour l'évolution des expéditions
    """
    try:
        annee_debut = int(request.GET.get('annee_debut', datetime.now().year))
        annee_fin = request.GET.get('annee_fin', None)
        
        if annee_fin:
            annee_fin = int(annee_fin)
        
        data = AnalyticsService.evolution_expeditions(annee_debut, annee_fin)
        data = convert_decimals_to_float(data)
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def api_evolution_ca(request):
    """
    API pour l'évolution du chiffre d'affaires
    """
    try:
        annee_debut = int(request.GET.get('annee_debut', datetime.now().year))
        annee_fin = request.GET.get('annee_fin', None)
        
        if annee_fin:
            annee_fin = int(annee_fin)
        
        data = AnalyticsService.evolution_chiffre_affaires(annee_debut, annee_fin)
        data = convert_decimals_to_float(data)
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def api_top_clients(request):
    """
    API pour les meilleurs clients
    """
    try:
        limite = int(request.GET.get('limite', 10))
        annee = request.GET.get('annee', None)
        
        if annee:
            annee = int(annee)
        
        data = AnalyticsService.top_clients(limite, annee)
        data = convert_decimals_to_float(data)
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def api_destinations_populaires(request):
    """
    API pour les destinations populaires
    """
    try:
        limite = int(request.GET.get('limite', 10))
        annee = request.GET.get('annee', None)
        
        if annee:
            annee = int(annee)
        
        data = AnalyticsService.destinations_populaires(limite, annee)
        data = convert_decimals_to_float(data)
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# ==================== ANALYSE OPÉRATIONNELLE ====================

@login_required
def analyse_operationnelle(request):
    """
    Page d'analyse opérationnelle détaillée
    """
    annee = int(request.GET.get('annee', datetime.now().year))
    annees_disponibles = list(Tournee.objects.dates('date_depart', 'year') .values_list('date_depart__year', flat=True)
        .distinct() .order_by('-date_depart__year') )
    
    if not annees_disponibles:
        annees_disponibles = range(2020, datetime.now().year + 1)
    
    context = { 'annee': annee, 'annees_disponibles': annees_disponibles, 'titre': 'Analyse Opérationnelle'}
    
    return render(request, 'analytics/operationnelle.html', context)


@login_required
@require_http_methods(["GET"])
def api_evolution_tournees(request):
    """
    API pour l'évolution des tournées
    """
    try:
        annee_debut = int(request.GET.get('annee_debut', datetime.now().year))
        annee_fin = request.GET.get('annee_fin', None)
        
        if annee_fin:
            annee_fin = int(annee_fin)
        
        data = AnalyticsService.evolution_tournees(annee_debut, annee_fin)
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def api_taux_reussite(request):
    """
    API pour le taux de réussite des livraisons
    """
    try:
        annee = request.GET.get('annee', None)
        
        if annee:
            annee = int(annee)
        
        data = AnalyticsService.taux_reussite_livraisons(annee)
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def api_top_chauffeurs(request):
    """
    API pour les meilleurs chauffeurs
    """
    try:
        limite = int(request.GET.get('limite', 10))
        annee = request.GET.get('annee', None)
        
        if annee:
            annee = int(annee)
        
        data = AnalyticsService.top_chauffeurs(limite, annee)
        data = convert_decimals_to_float(data)
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def api_zones_incidents(request):
    """
    API pour les zones avec incidents
    """
    try:
        limite = int(request.GET.get('limite', 10))
        annee = request.GET.get('annee', None)
        
        if annee:
            annee = int(annee)
        
        data = AnalyticsService.zones_incidents(limite, annee)
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def api_periodes_activite(request):
    """
    API pour les périodes de forte activité
    """
    try:
        annee = int(request.GET.get('annee', datetime.now().year))
        
        data = AnalyticsService.periodes_forte_activite(annee)
        data = convert_decimals_to_float(data)
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# ==================== STATISTIQUES ET KPI ====================

@login_required
def statistiques_kpi(request):
    """
    Page des statistiques et KPI
    """
    annee = int(request.GET.get('annee', datetime.now().year))
    annees_disponibles = list( Expedition.objects.dates('date_creation', 'year') .values_list('date_creation__year', flat=True) 
        .distinct() .order_by('-date_creation__year') )
    
    if not annees_disponibles:
        annees_disponibles = range(2020, datetime.now().year + 1)
    
    context = {'annee': annee, 'annees_disponibles': annees_disponibles, 'titre': 'Statistiques et KPI'}
    
    return render(request, 'analytics/kpi.html', context)


@login_required
@require_http_methods(["GET"])
def api_stats_generales(request):
    """
    API pour les statistiques générales
    """
    try:
        date_debut = request.GET.get('date_debut', None)
        date_fin = request.GET.get('date_fin', None)
        
        data = StatsService.statistiques_generales(date_debut, date_fin)
        data = convert_decimals_to_float(data)
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def api_kpi_expeditions(request):
    """
    API pour les KPI expéditions
    """
    try:
        annee = request.GET.get('annee', None)
        
        if annee:
            annee = int(annee)
        
        data = StatsService.kpi_expeditions(annee)
        data = convert_decimals_to_float(data)
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def api_kpi_financiers(request):
    """
    API pour les KPI financiers
    """
    try:
        annee = request.GET.get('annee', None)
        
        if annee:
            annee = int(annee)
        
        data = StatsService.kpi_financiers(annee)
        data = convert_decimals_to_float(data)
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def api_kpi_operationnels(request):
    """
    API pour les KPI opérationnels
    """
    try:
        annee = request.GET.get('annee', None)
        
        if annee:
            annee = int(annee)
        
        data = StatsService.kpi_operationnels(annee)
        data = convert_decimals_to_float(data)
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def api_kpi_qualite(request):
    """
    API pour les KPI qualité
    """
    try:
        annee = request.GET.get('annee', None)
        
        if annee:
            annee = int(annee)
        
        data = StatsService.kpi_qualite(annee)
        data = convert_decimals_to_float(data)
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# ==================== ANALYSES AVANCÉES ====================

@login_required
def analyses_avancees(request):
    """
    Page des analyses avancées
    """
    annee = int(request.GET.get('annee', datetime.now().year))
    annees_disponibles = list( Expedition.objects.dates('date_creation', 'year') .values_list('date_creation__year', flat=True)
        .distinct() .order_by('-date_creation__year') )
    
    if not annees_disponibles:
        annees_disponibles = range(2020, datetime.now().year + 1)
    
    context = { 'annee': annee, 'annees_disponibles': annees_disponibles, 'titre': 'Analyses Avancées' }
    
    return render(request, 'analytics/avancees.html', context)


@login_required
@require_http_methods(["GET"])
def api_comparaison_periodes(request):
    """
    API pour comparer deux périodes
    """
    try:
        date_debut1 = request.GET.get('date_debut1')
        date_fin1 = request.GET.get('date_fin1')
        date_debut2 = request.GET.get('date_debut2')
        date_fin2 = request.GET.get('date_fin2')
        data = StatsService.comparaison_periodes( date_debut1, date_fin1, date_debut2, date_fin2)
        data = convert_decimals_to_float(data)
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def api_saisonnalite(request):
    """
    API pour l'analyse de saisonnalité
    """
    try:
        annee = int(request.GET.get('annee', datetime.now().year))
        
        data = StatsService.analyse_saisonnalite(annee)
        data = convert_decimals_to_float(data)
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def api_rentabilite_destinations(request):
    """
    API pour la rentabilité par destination
    """
    try:
        annee = request.GET.get('annee', None)
        if annee: annee = int(annee)
        data = StatsService.analyse_rentabilite_destinations(annee)
        data = convert_decimals_to_float(data)
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


@login_required
@require_http_methods(["GET"])
def api_performance_vehicules(request):
    """
    API pour la performance des véhicules
    """
    try:
        annee = request.GET.get('annee', None)
        if annee: annee = int(annee)
        data = StatsService.analyse_performance_vehicules(annee)
        data = convert_decimals_to_float(data)
        
        return JsonResponse(data, safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=400)


# ==================== EXPORT PDF/EXCEL ====================

@login_required
def export_rapport_pdf(request):
    """
    Exporter le rapport en PDF
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.pdfgen import canvas
        
        annee = int(request.GET.get('annee', datetime.now().year))
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="rapport_analytics_{annee}.pdf"'
        p = canvas.Canvas(response, pagesize=A4)
        p.setFont("Helvetica-Bold", 16)
        p.drawString(100, 800, f"Rapport d'Analyse - Année {annee}")
        stats = StatsService.statistiques_generales()
        y = 750
        p.setFont("Helvetica", 12)
        p.drawString(100, y, f"Total Expéditions: {stats['total_expeditions']}")
        y -= 20
        p.drawString(100, y, f"Total Clients: {stats['total_clients']}")
        y -= 20
        p.drawString(100, y, f"CA Total: {float(stats['ca_total']):.2f} DA")
        p.showPage()
        p.save()
        
        return response
    except Exception as e:
        return HttpResponse(f"Erreur lors de la génération du PDF: {str(e)}", status=500)


@login_required
def export_rapport_excel(request):
    """
    Exporter le rapport en Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
        
        annee = int(request.GET.get('annee', datetime.now().year))
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Rapport {annee}"
        ws['A1'] = f"Rapport d'Analyse - Année {annee}"
        ws['A1'].font = Font(size=14, bold=True)
        stats = StatsService.statistiques_generales()
        
        row = 3
        ws[f'A{row}'] = "Statistiques Générales"
        ws[f'A{row}'].font = Font(bold=True)
        
        row += 1
        ws[f'A{row}'] = "Total Expéditions"
        ws[f'B{row}'] = stats['total_expeditions']
        
        row += 1
        ws[f'A{row}'] = "Total Clients"
        ws[f'B{row}'] = stats['total_clients']
        
        row += 1
        ws[f'A{row}'] = "CA Total"
        ws[f'B{row}'] = float(stats['ca_total'])

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="rapport_analytics_{annee}.xlsx"'
        
        wb.save(response)
        return response
    except Exception as e:
        return HttpResponse(f"Erreur lors de la génération du fichier Excel: {str(e)}", status=500)


# ==================== FONCTION UTILITAIRE ====================

def convert_decimals_to_float(obj):
    """
    Convertit récursivement tous les Decimal en float dans une structure de données
    """
    if isinstance(obj, Decimal):
        return float(obj)
    elif isinstance(obj, dict):
        return {key: convert_decimals_to_float(value) for key, value in obj.items()}
    elif isinstance(obj, list):
        return [convert_decimals_to_float(item) for item in obj]
    elif isinstance(obj, tuple):
        return tuple(convert_decimals_to_float(item) for item in obj)
    return obj

from django.http import HttpResponse

def emails_view(request):
    return HttpResponse("Page de gestion des emails et notifications")